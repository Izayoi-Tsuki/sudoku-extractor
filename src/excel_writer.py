"""
Excel 输出模块
功能：将数独数据输出到 Excel 文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Optional


class ExcelWriter:
    """数独 Excel 输出器"""
    
    def __init__(self, output_path: str = "sudoku_output.xlsx"):
        """
        初始化 Excel 写入器
        
        Args:
            output_path: 输出文件路径
        """
        self.output_path = output_path
        self.wb = Workbook()
        self.ws = None
    
    def _setup_styles(self):
        """设置样式"""
        # 粗体字体
        self.bold_font = Font(bold=True, size=12)
        
        # 居中对齐
        self.center_align = Alignment(
            horizontal='center',
            vertical='center'
        )
        
        # 边框
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 3x3 宫格的粗边框样式
        self.thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
    
    def create_sudoku_sheet(self, title: str = "Sudoku"):
        """
        创建数独工作表
        
        Args:
            title: 工作表标题
        """
        self.ws = self.wb.active
        self.ws.title = title
        self._setup_styles()
    
    def write_grid(self, grid: List[Optional[str]], filename: Optional[str] = None):
        """
        写入数独网格
        
        Args:
            grid: 81个元素的列表 (数字或空字符串)
            filename: 可选的输出文件名
        """
        output_path = filename or self.output_path
        
        if self.ws is None:
            self.create_sudoku_sheet()
        
        # 确保 grid 有81个元素
        if len(grid) < 81:
            grid = grid + [''] * (81 - len(grid))
        
        # 写入数据
        for row in range(9):
            for col in range(9):
                cell = self.ws.cell(
                    row=row + 1,
                    column=col + 1,
                    value=grid[row * 9 + col]
                )
                
                # 设置样式
                cell.border = self.thin_border
                cell.alignment = self.center_align
                cell.font = Font(size=14)
                
                # 设置 3x3 宫格边框
                if (col + 1) % 3 == 0 and col < 8:  # 右边框
                    cell.border = self.thick_border
                if (row + 1) % 3 == 0 and row < 8:  # 底边框
                    cell.border = self.thick_border
        
        # 设置列宽和行高
        for col in range(1, 10):
            self.ws.column_dimensions[get_column_letter(col)].width = 8
        for row in range(1, 10):
            self.ws.row_dimensions[row].height = 25
        
        # 添加标题
        self.ws.cell(row=11, column=1, value="Extracted from image").font = Font(italic=True)
        
        # 保存
        self.wb.save(output_path)
        print(f"✅ 已保存到: {output_path}")
    
    def write_with_metadata(self, grid: List[Optional[str]], 
                            source_file: str,
                            filename: Optional[str] = None):
        """
        写入数独网格（含元数据）
        
        Args:
            grid: 81个元素的列表
            source_file: 源图片文件名
            filename: 输出文件名
        """
        output_path = filename or self.output_path
        
        # 创建工作表
        self.create_sudoku_sheet("Sudoku")
        
        # 写入元数据
        self.ws.cell(row=1, column=1, value="Source File:").font = self.bold_font
        self.ws.cell(row=1, column=2, value=source_file)
        
        self.ws.cell(row=2, column=1, value="Extracted At:").font = self.bold_font
        from datetime import datetime
        self.ws.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # 空行
        self.ws.cell(row=4, column=1, value="")
        
        # 写入数独网格（从第6行开始）
        for row in range(9):
            for col in range(9):
                cell = self.ws.cell(
                    row=row + 6,
                    column=col + 1,
                    value=grid[row * 9 + col]
                )
                cell.border = self.thin_border
                cell.alignment = self.center_align
                cell.font = Font(size=14)
                
                # 3x3 边框
                if (col + 1) % 3 == 0 and col < 8:
                    cell.border = self.thick_border
                if (row + 1) % 3 == 0 and row < 8:
                    cell.border = self.thick_border
        
        # 设置列宽和行高
        for col in range(1, 10):
            self.ws.column_dimensions[get_column_letter(col)].width = 10
        for row in range(6, 15):
            self.ws.row_dimensions[row].height = 30
        
        # Save
        self.wb.save(output_path)
        print(f"[OK] Saved to: {output_path}")
    
    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()


if __name__ == "__main__":
    writer = ExcelWriter()
    print("✅ ExcelWriter 加载成功")
    print("   - 支持 write_grid(): 写入数独网格")
    print("   - 支持 write_with_metadata(): 写入含元数据的网格")
